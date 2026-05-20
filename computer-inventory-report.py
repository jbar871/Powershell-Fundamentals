#!/usr/bin/env python3
"""
Generate an HTML computer inventory report from the Freshservice export Excel file.
Reads the 'in' sheet.
Usage: python3 computer-inventory-report.py <input.xlsx> [output.html]
"""

import sys
import warnings
import html
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ── Column indices (0-based) for the 'in' sheet ─────────────────────────────
COL = {
    "asset_name":   1,   # B
    "ci_type":      2,   # C  CI Type (Desktop / Laptop)
    "product":      3,   # D  Product (model)
    "location":     4,   # E  Used By Location
    "email":        6,   # G  Used By Primary Email
    "name":         7,   # H  Used By Name
    "job_title":    8,   # I  Used By Job Title
    "user_active":  9,   # J  Used By Is Active (bool)
    "device_active":10,  # K  Device Active in AD
    "last_audit":   26,  # AA Last Audit Date
    "last_updated": 27,  # AB Last Updated Date
    "ad_changed":   28,  # AC AD whenChanged
    "ad_logon":     29,  # AD AD LastLogonDate
}

STALE_DAYS = 90

# ── Tech mapping: location code prefix → tech name ──────────────────────────
# Based on 2025 Field IT Escalation Chart v2 8/18/25
TECH_MAP = {
    # Jessie Barsky — Northeast + Atlantic (partial) + Southeast (partial)
    "8551": "Jessie Barsky",
    "8554": "Jessie Barsky",
    "8577": "Jessie Barsky",
    "8802": "Jessie Barsky",
    "8804": "Jessie Barsky",
    "8876": "Jessie Barsky",
    "8877": "Jessie Barsky",
    "8880": "Jessie Barsky",
    "9051": "Jessie Barsky",
    "9052": "Jessie Barsky",
    "9055": "Jessie Barsky",
    # Faisal Khan — Atlantic (partial) + Central + Ohio Valley
    "9201": "Faisal Khan",
    "9202": "Faisal Khan",
    "9203": "Faisal Khan",
    "9207": "Faisal Khan",
    "8378": "Faisal Khan",
    "8425": "Faisal Khan",
    "8603": "Faisal Khan",
    "8726": "Faisal Khan",
    "8927": "Faisal Khan",
    "8404": "Faisal Khan",
    "8450": "Faisal Khan",
    "8926": "Faisal Khan",
    "8928": "Faisal Khan",
    "8929": "Faisal Khan",
    "9001": "Faisal Khan",
    # Tony Justiniano — Southeast + Orlando Call Center
    "8253": "Tony Justiniano",
    "8259": "Tony Justiniano",
    "8277": "Tony Justiniano",
    "8278": "Tony Justiniano",
    "8281": "Tony Justiniano",
    "8282": "Tony Justiniano",
    "9062": "Tony Justiniano",
    "8298": "Tony Justiniano",
    # Murtaza Samma — Atlanta + Nashville + Texas + Houston Call Center
    "8302": "Murtaza Samma",
    "9103": "Murtaza Samma",
    "9126": "Murtaza Samma",
    "9127": "Murtaza Samma",
    "9128": "Murtaza Samma",
    "9131": "Murtaza Samma",
    "9132": "Murtaza Samma",
    "9133": "Murtaza Samma",
    "9134": "Murtaza Samma",
    "9138": "Murtaza Samma",
    "8299": "Murtaza Samma",
    # Joe Villalobos — West Division
    "8107": "Joe Villalobos",
    "8119": "Joe Villalobos",
    "8120": "Joe Villalobos",
    "8130": "Joe Villalobos",
    "8131": "Joe Villalobos",
    "8153": "Joe Villalobos",
    "8757": "Joe Villalobos",
    "8976": "Joe Villalobos",
    "9153": "Joe Villalobos",
    "9226": "Joe Villalobos",
    "9228": "Joe Villalobos",
    # Mario Rodriguez — Conway (Memphis) + Corporate
    "9106": "Mario Rodriguez",
    "9700": "Mario Rodriguez",
}

ALL_TECHS = sorted(set(TECH_MAP.values()))


def get_tech(location):
    if not location:
        return ""
    code = str(location).strip()[:4]
    return TECH_MAP.get(code, "")


def fmt_date(val):
    if val is None or val == "#N/A":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)


def days_ago(val, now):
    if not isinstance(val, datetime):
        return None
    return (now - val.replace(tzinfo=None)).days


def active_label(val):
    if val is True:
        return "Yes"
    if val is False:
        return "No"
    if val in ("#N/A", None):
        return "N/A"
    return str(val)


def row_class(user_active, device_active, audit_days, logon_days):
    if user_active is False:
        return "inactive-user"
    if user_active in ("#N/A", None):
        return "unknown-user"
    if device_active is False:
        return "inactive-device"
    if (audit_days is not None and audit_days > STALE_DAYS) or \
       (logon_days is not None and logon_days > STALE_DAYS):
        return "stale"
    return ""


def load_rows(path):
    wb = openpyxl.load_workbook(path)
    # find the 'in' sheet case-insensitively
    sheet_name = next((n for n in wb.sheetnames if n.lower() == "in"), None)
    if not sheet_name:
        print(f"ERROR: No 'in' sheet found. Available sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet_name]
    return list(ws.iter_rows(min_row=2, values_only=True))


def build_html(rows, source_file):
    now = datetime.now()
    generated = now.strftime("%Y-%m-%d %H:%M")

    total = len(rows)
    inactive_user  = sum(1 for r in rows if r[COL["user_active"]] is False)
    na_user        = sum(1 for r in rows if r[COL["user_active"]] in (None, "#N/A"))
    inactive_device = sum(1 for r in rows if r[COL["device_active"]] is False)
    stale = sum(
        1 for r in rows
        if r[COL["user_active"]] not in (False, None, "#N/A")
        and r[COL["device_active"]] is not False
        and (
            (days_ago(r[COL["last_audit"]], now) or 0) > STALE_DAYS
            or (days_ago(r[COL["ad_logon"]], now) or 0) > STALE_DAYS
        )
    )

    # Collect unique locations for the filter dropdown
    all_locations = sorted(
        {str(r[COL["location"]]) for r in rows if r[COL["location"]]},
        key=lambda x: x
    )
    location_options = "\n".join(
        f'      <option value="{html.escape(loc)}">{html.escape(loc)}</option>'
        for loc in all_locations
    )
    tech_options = "\n".join(
        f'      <option value="{html.escape(t)}">{html.escape(t)}</option>'
        for t in ALL_TECHS
    )

    table_rows = []
    for r in rows:
        ua         = r[COL["user_active"]]
        da         = r[COL["device_active"]]
        audit_days = days_ago(r[COL["last_audit"]], now)
        logon_days = days_ago(r[COL["ad_logon"]], now)
        cls        = row_class(ua, da, audit_days, logon_days)
        tech       = get_tech(r[COL["location"]])

        def d(col):
            return html.escape(fmt_date(r[COL[col]]))

        def s(col):
            v = r[COL[col]]
            return "" if v in (None, "#N/A") else html.escape(str(v))

        def stale_td(days, val):
            sc = ' class="stale-cell"' if (days is not None and days > STALE_DAYS) else ""
            return f"<td{sc}>{html.escape(fmt_date(val))}</td>"

        ua_cls = "yes" if ua is True else "no" if ua is False else "na"
        da_cls = "yes" if da is True else "no" if da is False else "na"

        table_rows.append(
            f'<tr class="{cls}">'
            f'<td>{s("asset_name")}</td>'
            f'<td>{s("ci_type")}</td>'
            f'<td>{s("product")}</td>'
            f'<td class="tech">{html.escape(tech)}</td>'
            f'<td>{s("location")}</td>'
            f'<td>{s("name")}</td>'
            f'<td>{s("email")}</td>'
            f'<td>{s("job_title")}</td>'
            f'<td class="bool {ua_cls}">{active_label(ua)}</td>'
            f'<td class="bool {da_cls}">{active_label(da)}</td>'
            + stale_td(audit_days, r[COL["last_audit"]])
            + f'<td>{d("last_updated")}</td>'
            + f'<td>{d("ad_changed")}</td>'
            + stale_td(logon_days, r[COL["ad_logon"]])
            + '</tr>'
        )

    rows_html = "\n".join(table_rows)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Computer Inventory Report</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         font-size: 13px; background: #f0f2f5; color: #1a1a2e; }}
  header {{ background: #1a1a2e; color: #fff; padding: 16px 24px;
            display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px; }}
  header h1 {{ font-size: 18px; font-weight: 600; }}
  header span {{ font-size: 11px; opacity: 0.6; }}

  .summary {{ display: flex; gap: 12px; padding: 16px 24px; flex-wrap: wrap; }}
  .card {{ background: #fff; border-radius: 8px; padding: 12px 18px;
           min-width: 130px; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  .card .num {{ font-size: 26px; font-weight: 700; }}
  .card .lbl {{ font-size: 11px; color: #666; margin-top: 2px; }}
  .card.red .num   {{ color: #d32f2f; }}
  .card.orange .num {{ color: #e65100; }}
  .card.yellow .num {{ color: #f9a825; }}
  .card.gray .num  {{ color: #555; }}

  .toolbar {{ padding: 0 24px 12px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }}
  .toolbar input[type=text] {{ padding: 7px 12px; border: 1px solid #ccc; border-radius: 6px;
                               font-size: 13px; width: 260px; }}
  .toolbar select {{ padding: 7px 10px; border: 1px solid #ccc; border-radius: 6px; font-size: 13px; }}
  .toolbar label  {{ font-size: 12px; color: #555; display: flex; align-items: center; gap: 4px; }}

  .wrap {{ padding: 0 24px 32px; overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff;
           border-radius: 8px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  th {{ background: #1a1a2e; color: #fff; padding: 10px; text-align: left;
        font-weight: 600; font-size: 12px; white-space: nowrap;
        cursor: pointer; user-select: none; position: sticky; top: 0; z-index: 2; }}
  th:hover {{ background: #2d2d4e; }}
  th .si {{ margin-left: 4px; opacity: 0.5; font-size: 10px; }}
  th.asc  .si::after {{ content: "▲"; }}
  th.desc .si::after {{ content: "▼"; }}
  th:not(.asc):not(.desc) .si::after {{ content: "⇅"; }}

  td {{ padding: 8px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: middle; white-space: nowrap; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: rgba(0,0,0,.025); }}

  tr.inactive-user   td {{ background: #fdecea; }}
  tr.inactive-user:hover td {{ background: #fcd8d5; }}
  tr.unknown-user    td {{ background: #fff8e1; }}
  tr.unknown-user:hover td {{ background: #fff0c0; }}
  tr.inactive-device td {{ background: #fff3e0; }}
  tr.stale           td {{ background: #f3f8ff; }}

  td.bool {{ text-align: center; font-weight: 600; font-size: 12px; }}
  td.bool.yes {{ color: #2e7d32; }}
  td.bool.no  {{ color: #c62828; }}
  td.bool.na  {{ color: #888; }}
  td.stale-cell {{ color: #e65100; font-weight: 600; }}
  td.tech {{ font-weight: 600; font-size: 12px; color: #1a1a2e; }}

  .legend {{ display: flex; gap: 12px; padding: 0 24px 16px; flex-wrap: wrap; font-size: 11px; }}
  .legend span {{ display: flex; align-items: center; gap: 5px; }}
  .sw {{ width: 12px; height: 12px; border-radius: 2px; display: inline-block; flex-shrink: 0; }}
  #row-count {{ font-size: 12px; color: #555; padding: 0 24px 8px; }}
</style>
</head>
<body>
<header>
  <h1>Computer Inventory Report</h1>
  <span>Source: {html.escape(source_file)} &nbsp;|&nbsp; Generated: {generated}</span>
</header>

<div class="summary">
  <div class="card gray">  <div class="num">{total}</div>          <div class="lbl">Total Devices</div></div>
  <div class="card red">   <div class="num">{inactive_user}</div>  <div class="lbl">Inactive User</div></div>
  <div class="card yellow"><div class="num">{na_user}</div>        <div class="lbl">User Unknown / N/A</div></div>
  <div class="card orange"><div class="num">{inactive_device}</div><div class="lbl">Inactive in AD</div></div>
  <div class="card orange"><div class="num">{stale}</div>          <div class="lbl">Stale (&gt;{STALE_DAYS}d)</div></div>
</div>

<div class="toolbar">
  <input type="text" id="search" placeholder="Search asset, user, email…">
  <label>Type:
    <select id="filter-type">
      <option value="">All Types</option>
      <option value="Desktop">Desktop</option>
      <option value="Laptop">Laptop</option>
    </select>
  </label>
  <label>Tech:
    <select id="filter-tech">
      <option value="">All Techs</option>
{tech_options}
    </select>
  </label>
  <label>Location:
    <select id="filter-loc">
      <option value="">All Locations</option>
{location_options}
    </select>
  </label>
  <label>User Active:
    <select id="filter-user">
      <option value="">All</option>
      <option value="Yes">Yes</option>
      <option value="No">No</option>
      <option value="N/A">N/A</option>
    </select>
  </label>
  <label>Device Active in AD:
    <select id="filter-device">
      <option value="">All</option>
      <option value="Yes">Yes</option>
      <option value="No">No</option>
    </select>
  </label>
  <label><input type="checkbox" id="filter-stale"> Stale only (&gt;{STALE_DAYS}d)</label>
</div>
<div id="row-count"></div>

<div class="legend">
  <span><span class="sw" style="background:#fdecea"></span> Inactive User</span>
  <span><span class="sw" style="background:#fff8e1"></span> User Unknown/N/A</span>
  <span><span class="sw" style="background:#fff3e0"></span> Inactive Device in AD</span>
  <span><span class="sw" style="background:#f3f8ff"></span> Stale (&gt;{STALE_DAYS}d)</span>
  <span style="color:#e65100;font-weight:600">Orange date = stale</span>
</div>

<div class="wrap">
<table id="tbl">
  <thead><tr>
    <th data-col="0">Asset Name<span class="si"></span></th>
    <th data-col="1">Type<span class="si"></span></th>
    <th data-col="2">Model<span class="si"></span></th>
    <th data-col="3">Tech<span class="si"></span></th>
    <th data-col="4">Location<span class="si"></span></th>
    <th data-col="5">Used By Name<span class="si"></span></th>
    <th data-col="6">Email<span class="si"></span></th>
    <th data-col="7">Job Title<span class="si"></span></th>
    <th data-col="8">User Active<span class="si"></span></th>
    <th data-col="9">Device Active in AD<span class="si"></span></th>
    <th data-col="10">Last Audit Date<span class="si"></span></th>
    <th data-col="11">Last Updated Date<span class="si"></span></th>
    <th data-col="12">AD whenChanged<span class="si"></span></th>
    <th data-col="13">AD LastLogonDate<span class="si"></span></th>
  </tr></thead>
  <tbody id="tbody">
{rows_html}
  </tbody>
</table>
</div>

<script>
const tbody     = document.getElementById("tbody");
const allRows   = Array.from(tbody.querySelectorAll("tr"));
const search    = document.getElementById("search");
const fType     = document.getElementById("filter-type");
const fTech     = document.getElementById("filter-tech");
const fLoc      = document.getElementById("filter-loc");
const fUser     = document.getElementById("filter-user");
const fDevice   = document.getElementById("filter-device");
const fStale    = document.getElementById("filter-stale");
const rowCount  = document.getElementById("row-count");

// col indices: 0=Asset 1=Type 2=Model 3=Tech 4=Location 5=Name 6=Email 7=JobTitle 8=UserActive 9=DeviceActive
function applyFilters() {{
  const q      = search.value.toLowerCase();
  const type   = fType.value;
  const tech   = fTech.value;
  const loc    = fLoc.value;
  const ua     = fUser.value;
  const da     = fDevice.value;
  const stOnly = fStale.checked;
  let vis = 0;

  allRows.forEach(row => {{
    const cells = row.querySelectorAll("td");
    const matchQ      = !q      || row.textContent.toLowerCase().includes(q);
    const matchType   = !type   || cells[1]?.textContent.trim() === type;
    const matchTech   = !tech   || cells[3]?.textContent.trim() === tech;
    const matchLoc    = !loc    || cells[4]?.textContent.trim() === loc;
    const matchUA     = !ua     || cells[8]?.textContent.trim() === ua;
    const matchDA     = !da     || cells[9]?.textContent.trim() === da;
    const matchStale  = !stOnly || row.className === "stale";

    const show = matchQ && matchType && matchTech && matchLoc && matchUA && matchDA && matchStale;
    row.style.display = show ? "" : "none";
    if (show) vis++;
  }});

  rowCount.textContent = `Showing ${{vis}} of ${{allRows.length}} devices`;
}}

[search, fType, fTech, fLoc, fUser, fDevice, fStale].forEach(el =>
  el.addEventListener(el.type === "checkbox" ? "change" : "input", applyFilters)
);
applyFilters();

// Column sorting
let sortCol = -1, sortDir = 1;
document.querySelectorAll("th[data-col]").forEach(th => {{
  th.addEventListener("click", () => {{
    const col = parseInt(th.dataset.col);
    sortDir = (sortCol === col) ? -sortDir : 1;
    sortCol = col;
    document.querySelectorAll("th").forEach(h => h.classList.remove("asc","desc"));
    th.classList.add(sortDir === 1 ? "asc" : "desc");
    const sorted = allRows.slice().sort((a, b) => {{
      const av = a.querySelectorAll("td")[col]?.textContent.trim() || "";
      const bv = b.querySelectorAll("td")[col]?.textContent.trim() || "";
      return av.localeCompare(bv, undefined, {{numeric: true}}) * sortDir;
    }});
    sorted.forEach(r => tbody.appendChild(r));
    applyFilters();
  }});
}});
</script>
</body>
</html>"""


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 computer-inventory-report.py <input.xlsx> [output.html]")
        sys.exit(1)

    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else Path(src).stem + "-report.html"

    print(f"Reading {src} (sheet: 'in')...")
    rows = load_rows(src)
    print(f"  {len(rows)} devices loaded.")

    print("Building report...")
    content = build_html(rows, Path(src).name)

    with open(out, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"  Saved: {out}")


if __name__ == "__main__":
    main()
